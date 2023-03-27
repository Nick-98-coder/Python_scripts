import os
import shutil
import openpyxl as op

#enter excel sheet path here
ptr2 = op.load_workbook('C:/Users/Tnluser/Documents/File_path.xlsx')
dataframe1 = ptr2.active

#file_finder_function
def search_and_copy(file_name, src_path, dest_path):
    for root, dirs, files in os.walk(src_path):
        for file in files:
            if file == file_name:
                src_file = os.path.join(root, file)
                shutil.copy(src_file, dest_path)
                print(f'{file_name} copied to {dest_path}')
                return
    print(f'{file_name} not found in {src_path}')

#fetching file path from excel sheet.

for col in range(0,1):
    for row in dataframe1.iter_rows(2,dataframe1.max_row):
            file_name = row[col].value
            print("File name", file_name)
for col in range(1,2):
    for row in dataframe1.iter_rows(2,dataframe1.max_row):
            src_path = row[col].value
            print("From Path", src_path)
for col in range(2,3):
    for row in dataframe1.iter_rows(2,dataframe1.max_row):
            dest_path = row[col].value
            print("To_path", dest_path) 

           
for i in range(6):
    search_and_copy(file_name[i], src_path[i], dest_path[i])

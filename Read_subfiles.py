import openpyxl
import os

def process_files(folder_path, sheet, row):
    files = os.listdir(folder_path)
    for file in files:
        file_path = os.path.join(folder_path, file)
        if os.path.isfile(file_path):
            filename, ext = os.path.splitext(file)
            present_with_ext = filename + ext
            print("Filename:", present_with_ext)
            presentfile_cell = sheet.cell(row=row, column=1)
            presentfile_cell.value = present_with_ext
            folder_path_cell = sheet.cell(row=row, column=2)
            folder_path_cell.value = folder_path
            row += 1
        elif os.path.isdir(file_path):
            row = process_files(file_path, sheet, row)
    return row

# Read the Excel file
wb = openpyxl.load_workbook(r'C:\Users\nikhil.srinivas\Documents\filename.xlsx')
sheet = wb.active

# Prompt the user for the folder path
folder_path = input("Enter your path:\n")

# Start processing files recursively
start_row = 2  # Start from row 2 in the Excel sheet
process_files(folder_path, sheet, start_row)

# Save the updated Excel file
wb.save(r'C:\Users\nikhil.srinivas\Documents\filename.xlsx')

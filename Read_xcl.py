import openpyxl as op

ptr2 = op.load_workbook('C:/Users/Tnluser/Documents/File_path.xlsx')
dataframe1 = ptr2.active


for col in range(0,1):
    for row in dataframe1.iter_rows(2,dataframe1.max_row):
        File_name = row[col].value
        print("File name", File_name)
for col in range(1,2):
    for row in dataframe1.iter_rows(2,dataframe1.max_row):
        From_path = row[col].value
        print("From Path", From_path)
for col in range(2,3):
    for row in dataframe1.iter_rows(2,dataframe1.max_row):
        To_path = row[col].value
        print("To_path", To_path)     

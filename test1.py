import openpyxl 
import os

# read the Excel file
wb = openpyxl.load_workbook('C:/Users/Tnluser/Documents/file_rename.xlsx')
sheet = wb.active

folder_path = "C:/Users/Tnluser/Documents/Search"
files = os.listdir(folder_path)
n = len(files)
print(n)

for i, file in zip(range(2, n+1), files):
    filename, ext = os.path.splitext(file)
    old_name_with_ext = filename + ext
    print("oldname: ", old_name_with_ext)

    # Enter old_name_with_ext value to the excel sheet
    for i, value in zip(range(2,n+1),old_name_with_ext):
        cell = sheet.cell(row=i+2, column=3)
        cell.value = value
        print(cell.value)

    ext = os.path.splitext(file)[1]
    new_name = sheet['B' + str(i)].value
    new_name_with_ext = new_name + ext
    for i, value in zip(range(2,n+1),new_name_with_ext):
        cell = sheet.cell(row=i+2, column=5)
        cell.value = value
        print(cell.value)
    print("new name:", new_name_with_ext)

# save the changes
#wb.save("C:/Users/Tnluser/Documents/file_rename.xlsx")

"""
# rename the file using the os module
os.rename(old_name_with_ext, new_name_with_ext)
"""

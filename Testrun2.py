import os
import openpyxl

# load the Excel workbook
wb = openpyxl.load_workbook('C:/Users/Tnluser/Documents/file_rename.xlsx')
sheet = wb.active

# get the number of rows in the sheet
n = 3
net_path = str(input("enter your path:\n"))

# iterate through the rows of the sheet use n+1 in (2, 5)
file_names = os.listdir(net_path)
x=len(file_names)
print(x)

for j, value in zip(range(x), file_names): 
    Present_cell = sheet.cell(row=j+2, column=1)
    Present_cell.value = value
    Present_files = value
    print("Present_files", Present_files)
    Present_WExt, ext = os.path.splitext(value)
    Present_ext_cell = sheet.cell(row=j+2, column=2)
    Present_ext_cell.value = Present_WExt
    print("Present_WExt", Present_WExt)
    old_name = str(sheet['C' + str(j+2)].value)
    print("old_name", old_name)
    new_name = str(sheet['D' + str(j+2)].value)
    print("new_name", new_name)
    if Present_files in file_names:
        # linear search algorithm to find the index of the old file name
        index = None
        for i in range(len(file_names)):
            if file_names[i] == old_name:
                index = i
                break
        if index is not None:
            # rename the file
            new_filename = os.path.join(net_path, new_name + ext)
            os.rename(os.path.join(net_path, file_names[index]), new_filename)
            new_gen_name = "Done"
            print("new_gen_name", new_filename)
            New_gen_cell = sheet.cell(row=j+2, column=5)
            New_gen_cell.value = new_gen_name
            print(new_gen_name)
        else:
            print("Failed to find old file name:", old_name)
    else:
        print("Old file name not found:", old_name)

# save the workbook
wb.save('C:/Users/Tnluser/Documents/file_rename.xlsx')
